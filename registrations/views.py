from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import IntegrityError
from django.core.mail import send_mail
from django.conf import settings

from events.models import Event
from .models import Registration
from .forms import RegistrationForm
from django.db.models import Q
from django.http import HttpResponse
import openpyxl


def register_event(request, event_id):

    event = get_object_or_404(Event, id=event_id)

    if not event.status:
        messages.error(request, "Registration Closed.")
        return redirect("event_detail", id=event.id)

    if event.available_seats <= 0:
        messages.error(request, "No seats available.")
        return redirect("event_detail", id=event.id)

    if request.method == "POST":

        form = RegistrationForm(request.POST)

        if form.is_valid():

            email = form.cleaned_data["email"].strip().lower()

            if Registration.objects.filter(
                event=event,
                email__iexact=email
            ).exists():

                messages.warning(
                    request,
                    "You have already registered for this event."
                )

                return redirect("event_detail", id=event.id)

            registration = form.save(commit=False)
            registration.event = event
            registration.email = email

            try:
                registration.save()

                event.available_seats -= 1
                event.save()

                send_mail(

                    subject="Event Registration Successful - EventNest",

                    message=f"""

                Hello {registration.full_name},

                Congratulations!

                Your registration has been completed successfully.

                Event Details

                Event Name : {event.title}

                Organizer : {event.organizer}

                Venue : {event.venue}

                Event Date : {event.event_date}

                Start Time : {event.start_time}

                End Time : {event.end_time}

                Status : {registration.status}

                Thank you for registering.

                Regards,
                EventNest Team

                """,

                    from_email=settings.DEFAULT_FROM_EMAIL,

                    recipient_list=[registration.email],

                    fail_silently=False,

                )

                messages.success(
                    request,
                    "Registration Successful."
                )


            except Exception as e:

                print("EMAIL ERROR :", e)

                raise

            return redirect("event_detail", id=event.id)

    else:

        form = RegistrationForm()

    return render(
        request,
        "registrations/register_event.html",
        {
            "form": form,
            "event": event
        }
    )

def cancel_registration(request, id):

    registration = get_object_or_404(
        Registration,
        id=id
    )

    if request.method == "POST":

        if registration.status != "Cancelled":

            registration.status = "Cancelled"
            registration.save()

            event = registration.event
            event.available_seats += 1
            event.save()

            messages.success(
                request,
                "Registration Cancelled Successfully."
            )

        return redirect("registration_list")

    return render(
        request,
        "registrations/cancel_registration.html",
        {
            "registration": registration
        }
    )



def registration_list(request):

    query = request.GET.get("q", "")

    registrations = Registration.objects.select_related("event")

    if query:

        registrations = registrations.filter(

            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(event__title__icontains=query)

        )

    return render(

        request,

        "registrations/registration_list.html",

        {

            "registrations": registrations,

            "query": query

        }

    )


def registration_detail(request, id):

    registration = get_object_or_404(

        Registration,

        id=id

    )

    return render(

        request,

        "registrations/registration_detail.html",

        {

            "registration": registration

        }

    )

def update_registration_status(request, id, status):

    registration = get_object_or_404(
        Registration,
        id=id
    )

    if status in ["Pending", "Confirmed", "Cancelled"]:

        if registration.status != "Cancelled" and status == "Cancelled":

            registration.event.available_seats += 1
            registration.event.save()

        elif registration.status == "Cancelled" and status != "Cancelled":

            if registration.event.available_seats > 0:
                registration.event.available_seats -= 1
                registration.event.save()

        registration.status = status
        registration.save()

        messages.success(
            request,
            "Registration Status Updated Successfully."
        )

    return redirect("registration_detail", id=registration.id)


from django.utils.timezone import now


def export_registrations_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Registrations"

    headers = [

        "ID",
        "Student Name",
        "Email",
        "Phone",
        "College",
        "Department",
        "Year",
        "Event",
        "Status",
        "Registration Date"

    ]

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=1,
            column=column
        )

        cell.value = header

        cell.font = openpyxl.styles.Font(
            bold=True
        )

    registrations = Registration.objects.select_related(
        "event"
    ).all()

    row = 2

    for registration in registrations:

        sheet.cell(row=row, column=1).value = registration.id
        sheet.cell(row=row, column=2).value = registration.full_name
        sheet.cell(row=row, column=3).value = registration.email
        sheet.cell(row=row, column=4).value = registration.phone
        sheet.cell(row=row, column=5).value = registration.college
        sheet.cell(row=row, column=6).value = registration.department
        sheet.cell(row=row, column=7).value = registration.year
        sheet.cell(row=row, column=8).value = registration.event.title
        sheet.cell(row=row, column=9).value = registration.status
        sheet.cell(
            row=row,
            column=10
        ).value = registration.registration_date.strftime(
            "%d-%m-%Y %H:%M"
        )

        row += 1

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    filename = now().strftime(
        "Registrations_%d_%m_%Y.xlsx"
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response